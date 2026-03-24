import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import io

# 页面基础配置（适配手机端）
st.set_page_config(
    page_title="Excel图片批量插入工具",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# 页面标题与说明
st.title("📊 Excel图片批量插入工具")
st.markdown("""
本工具专为批量插入图片到Excel设计，完美兼容Excel2016及以上版本，核心特性：
- 手机端直接上传相册图片，按选择顺序自动排列
- 图片自适应单元格大小，**随单元格缩放同步变化**
- 支持新建空白Excel或上传已有Excel文件
- 自定义插入起始位置、排列方向，实时预览插入效果
- 生成后可直接导出下载，无需本地安装任何软件
""")
st.divider()

# 1. 选择Excel源文件
st.subheader("1. 选择Excel源文件")
excel_source = st.radio(
    "Excel来源",
    ["新建空白Excel", "上传已有Excel文件"],
    horizontal=True
)

wb = None
ws = None

if excel_source == "上传已有Excel文件":
    uploaded_excel = st.file_uploader(
        "上传Excel文件（仅支持.xlsx格式）",
        type=["xlsx"],
        help="仅支持Excel2016及以上版本的.xlsx文件，不支持旧版.xls格式"
    )
    if uploaded_excel:
        try:
            wb = load_workbook(uploaded_excel)
            ws = wb.active
            st.success(f"✅ 已加载Excel文件，当前激活工作表：{ws.title}")
        except Exception as e:
            st.error(f"❌ 加载Excel失败：{str(e)}，请检查文件格式是否正确、是否损坏")
else:
    wb = Workbook()
    ws = wb.active
    st.info("ℹ️ 已创建空白Excel文件，默认工作表：Sheet")

st.divider()

# 2. 上传图片
st.subheader("2. 上传图片")
uploaded_images = st.file_uploader(
    "选择图片（支持多选，按选择顺序排列）",
    type=["png", "jpg", "jpeg", "bmp", "gif"],
    accept_multiple_files=True,
    help="电脑端按住Ctrl/Command多选，手机端可直接从相册批量选择"
)

image_count = len(uploaded_images)
if image_count > 0:
    st.success(f"✅ 已上传 {image_count} 张图片")
else:
    st.warning("⚠️ 请上传需要插入的图片")

st.divider()

# 3. 设置插入参数
st.subheader("3. 设置插入参数")
col1, col2 = st.columns(2)
with col1:
    start_row = st.number_input(
        "起始行号",
        min_value=1,
        value=1,
        step=1,
        help="Excel行号，从1开始，例如从第2行开始就填2"
    )
    arrange_direction = st.radio(
        "图片排列方向",
        ["纵向（向下排列）", "横向（向右排列）"],
        horizontal=True
    )
with col2:
    start_col = st.number_input(
        "起始列号",
        min_value=1,
        value=1,
        step=1,
        help="Excel列号，从1开始，例如B列就填2"
    )
    margin_percent = st.slider(
        "单元格内边距（%）",
        min_value=0,
        max_value=20,
        value=5,
        step=1,
        help="图片与单元格边框的留白，避免图片顶满单元格"
    )

st.divider()

# 4. 插入位置预览
if image_count > 0:
    st.subheader("4. 插入位置预览")
    preview_container = st.container()
    with preview_container:
        for i, img_file in enumerate(uploaded_images):
            # 计算目标单元格位置
            if arrange_direction == "纵向（向下排列）":
                target_row = start_row + i
                target_col = start_col
            else:
                target_row = start_row
                target_col = start_col + i
            cell_addr = f"{get_column_letter(target_col)}{target_row}"
            
            # 生成缩略图预览
            img_file.seek(0)
            pil_img = PILImage.open(img_file)
            pil_img.thumbnail((120, 120))
            buf = io.BytesIO()
            pil_img.save(buf, format="PNG")
            
            # 展示预览
            preview_col1, preview_col2 = st.columns([1, 5])
            with preview_col1:
                st.markdown(f"**序号{i+1}**\n\n{cell_addr}")
            with preview_col2:
                st.image(buf.getvalue(), width=100)
            st.divider()

# 5. 生成并导出Excel
st.subheader("5. 生成并导出Excel")
generate_btn = st.button("生成Excel文件", type="primary", use_container_width=True)

if generate_btn:
    if not wb or not ws:
        st.error("❌ 请先选择或创建Excel文件")
    elif image_count == 0:
        st.error("❌ 请先上传需要插入的图片")
    else:
        try:
            # 遍历图片，批量插入Excel
            for i, img_file in enumerate(uploaded_images):
                # 计算目标单元格
                if arrange_direction == "纵向（向下排列）":
                    target_row = start_row + i
                    target_col = start_col
                else:
                    target_row = start_row
                    target_col = start_col + i
                
                # 转换单元格宽高为像素（适配Excel单位）
                col_letter = get_column_letter(target_col)
                # 列宽转换：1字符单位 ≈ 8像素，默认列宽8.43
                col_width = ws.column_dimensions[col_letter].width if col_letter in ws.column_dimensions else 8.43
                cell_width_px = col_width * 8
                # 行高转换：1磅 ≈ 1.333像素，默认行高15
                row_height = ws.row_dimensions[target_row].height if target_row in ws.row_dimensions else 15
                cell_height_px = row_height * 1.333
                
                # 计算边距后的可用空间
                margin_scale = 1 - margin_percent / 100
                available_width = cell_width_px * margin_scale
                available_height = cell_height_px * margin_scale
                
                # 图片等比例缩放（不拉伸变形）
                img_file.seek(0)
                pil_img = PILImage.open(img_file)
                img_w, img_h = pil_img.size
                scale = min(available_width / img_w, available_height / img_h)
                resized_img = pil_img.resize((int(img_w * scale), int(img_h * scale)), PILImage.Resampling.LANCZOS)
                
                # 转换为Openpyxl支持的图片格式
                img_buf = io.BytesIO()
                resized_img.save(img_buf, format="PNG")
                img_buf.seek(0)
                xl_image = OpenpyxlImage(img_buf)
                
                # 核心：双单元格锚定，实现随单元格缩放
                anchor = TwoCellAnchor()
                # Openpyxl锚定从0开始计数，Excel行号从1开始，需减1
                anchor.from_row = target_row - 1
                anchor.from_col = target_col - 1
                anchor.to_row = target_row
                anchor.to_col = target_col
                # 关键配置：editAs="twoCell" 实现图片随单元格同步缩放
                anchor.editAs = "twoCell"
                anchor.add_pic(xl_image)
                ws.drawing.add(anchor)
            
            # 保存Excel到内存
            excel_buf = io.BytesIO()
            wb.save(excel_buf)
            excel_buf.seek(0)
            
            st.success("✅ Excel文件生成成功！点击下方按钮下载")
            # 导出下载按钮
            st.download_button(
                label="📥 下载Excel文件",
                data=excel_buf,
                file_name="批量插入图片的表格.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        except Exception as e:
            st.error(f"❌ 生成Excel失败：{str(e)}，请检查参数设置")

# 底部使用说明
st.divider()
st.markdown("""
### 核心使用说明
1. **手机端使用**：直接在浏览器打开部署后的链接，即可从手机相册选择图片上传，无需安装APP
2. **随单元格缩放**：生成的Excel中，图片已绑定单元格，拖动调整单元格宽高时，图片会自动同步缩放
3. **顺序保证**：严格按照你选择图片的顺序依次插入，不会打乱排序
4. **兼容性**：生成的.xlsx文件完美兼容Excel2016、2019、365及WPS表格
5. **注意事项**：仅支持.xlsx格式，旧版.xls文件请先另存为.xlsx格式后再上传
""")
